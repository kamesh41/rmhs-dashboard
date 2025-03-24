import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.db.models import Sum
from datetime import datetime
import pytz
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

def export_to_excel(queryset, filename, columns=None, title=None, material_summary=None, operation_summary=None, summary_by_status=None):
    """Export queryset to Excel file with summary sections"""
    if columns is None:
        columns = [field.name for field in queryset.model._meta.fields]
    
    # Check if this is a Rake report
    is_rake_report = 'rake_id' in columns if columns else False
    
    # Convert queryset to list of dictionaries
    data = []
    for obj in queryset:
        row = {}
        for col in columns:
            if col in ['material', 'rake_material']:
                # For ForeignKey fields, get the name
                value = getattr(obj, col)
                row[col] = str(value) if value else ''
            else:
                # For regular fields
                if hasattr(obj, col):
                    value = getattr(obj, col)
                    # Format datetime fields
                    if isinstance(value, datetime):
                        # Convert to Asia/Kolkata timezone to match what's displayed in the UI
                        if value.tzinfo is None:
                            # If naive datetime, assume it's in UTC and convert to IST
                            ist = pytz.timezone('Asia/Kolkata')
                            value = pytz.utc.localize(value).astimezone(ist)
                        else:
                            # If already timezone aware, just convert to IST
                            ist = pytz.timezone('Asia/Kolkata')
                            value = value.astimezone(ist)
                        # Preserve exact time as stored - no timezone conversion
                        row[col] = value.strftime('%d-%b-%Y %H:%M')
                    else:
                        row[col] = value
                else:
                    # Check if it's using the new 'quantity' attribute name instead of 'tonnage'
                    if col == 'tonnage' and hasattr(obj, 'quantity'):
                        row[col] = getattr(obj, 'quantity')
                    else:
                        row[col] = ''
        data.append(row)
    
    # Create pandas DataFrame from data
    df = pd.DataFrame(data)
    
    # Create an in-memory Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add title as a separate worksheet if provided
        if title:
            workbook = writer.book
            title_sheet = workbook.create_sheet("Report Info", 0)
            title_sheet['A1'] = title
            title_sheet['A2'] = f"Generated on: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
            
            # Set column width
            title_sheet.column_dimensions['A'].width = 50
            
            # Style the title
            title_cell = title_sheet['A1']
            title_cell.font = Font(size=14, bold=True)
            
            # Style the date
            date_cell = title_sheet['A2']
            date_cell.font = Font(size=10, italic=True)
        
        # Main data sheet
        df.to_excel(writer, index=False, sheet_name='Data')
        
        # Format the worksheet
        worksheet = writer.sheets['Data']
        for idx, col in enumerate(df.columns, 1):
            # Set column width based on content
            column_width = max(
                len(str(col)), 
                df[col].astype(str).map(len).max() if not df.empty else 0
            )
            column_width = min(max(column_width + 2, 10), 50)  # Min 10, Max 50
            column_letter = get_column_letter(idx)
            worksheet.column_dimensions[column_letter].width = column_width
            
            # Set header style
            header_cell = worksheet[f"{column_letter}1"]
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
        # Add Status Summary sheet
        if summary_by_status is not None:
            status_df = pd.DataFrame([{
                'Status': item['rake_status'],
                'Count': item['total']
            } for item in summary_by_status])
            
            if not status_df.empty:
                status_df.to_excel(writer, index=False, sheet_name='Status Summary')
        
        # Add Material Summary sheet for Rake reports
        if is_rake_report:
            # Add Material Summary sheet for Rake reports if material summary exists
            if material_summary is not None:
                material_df = pd.DataFrame([{
                    'Material': item['rake_material'],
                    'Count': item['count']
                } for item in material_summary])
                
                if not material_df.empty:
                    material_df.to_excel(writer, index=False, sheet_name='Material Summary')
        
        # Add Material Summary sheet for Operation reports
        elif material_summary is not None:
            # Create a list to store the material data
            material_data = []
            
            for item in material_summary:
                material_row = {
                    'Material': item['material__name'],
                    'Total Quantity (Tons)': item.get('total', 0)
                }
                material_data.append(material_row)
            
            if material_data:
                material_df = pd.DataFrame(material_data)
                material_df.to_excel(writer, index=False, sheet_name='Material Summary')
        
        # Add Operation Summary sheet for Operation reports
        if operation_summary is not None and not is_rake_report:
            # Check if operation_summary is a dictionary (new format) or list (old format)
            if isinstance(operation_summary, dict):
                # Create a list to store the operation data
                operation_data = []
                
                for op_type, total in operation_summary.items():
                    operation_row = {
                        'Operation Type': op_type.title(),
                        'Total Quantity (Tons)': total
                    }
                    operation_data.append(operation_row)
                
                if operation_data:
                    operation_df = pd.DataFrame(operation_data)
                    operation_df.to_excel(writer, index=False, sheet_name='Operation Summary')
            else:
                # Legacy format (list of dicts)
                operation_df = pd.DataFrame([{
                    'Operation Type': item['operation_type'],
                    'Total Quantity (Tons)': item.get('total', 0)
                } for item in operation_summary])
                
                if not operation_df.empty:
                    operation_df.to_excel(writer, index=False, sheet_name='Operation Summary')
    
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    return response

def export_to_pdf(queryset, filename, columns, title=None, material_summary=None, operation_summary=None):
    """
    Export queryset data to PDF
    
    Args:
        queryset: Django queryset to export
        filename: Base filename for the PDF
        columns: List of column names to include
        title: Optional title for the report
        material_summary: Optional material summary data
        operation_summary: Optional operation summary data
    
    Returns:
        HttpResponse with PDF attachment
    """
    # Configure response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    # Create PDF object
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), title=filename)
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Create document elements
    elements = []
    
    # Add title
    if title:
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))
    
    # Add date
    date_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    elements.append(Paragraph(date_text, normal_style))
    elements.append(Spacer(1, 12))
    
    # Check if this is a Rake report based on column names
    is_rake_report = 'rake_id' in columns if columns else False
    
    # Add Material Summary Section if provided
    if material_summary:
        elements.append(Paragraph("Material Summary", heading_style))
        elements.append(Spacer(1, 10))
        
        # Prepare material summary data
        if is_rake_report:
            # For rake reports, material summary shows counts
            mat_data = [['Material', 'Count']]
            for item in material_summary:
                mat_data.append([
                    item['rake_material'], 
                    str(item['count'])
                ])
        else:
            # For operation reports, material summary shows quantities
            mat_data = [['Material', 'Total Quantity (Tons)']]
            for item in material_summary:
                mat_data.append([
                    item['material__name'], 
                    f"{item['total']:.2f}"
                ])
        
        # Create material summary table
        mat_table = Table(mat_data, colWidths=[3*inch, 2*inch])
        mat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(mat_table)
        elements.append(Spacer(1, 20))
    
    # Add Operation Summary Section if provided (for operations, not rakes)
    if operation_summary and not is_rake_report:
        elements.append(Paragraph("Operation Summary", heading_style))
        elements.append(Spacer(1, 10))
        
        # Prepare operation summary data
        op_data = [['Operation Type', 'Total Quantity (Tons)']]
        
        # Handle operation_summary as a dictionary
        if isinstance(operation_summary, dict):
            for op_type, total in operation_summary.items():
                op_data.append([
                    op_type.title(), 
                    f"{total:.2f}"
                ])
        
        # Create operation summary table
        op_table = Table(op_data, colWidths=[3*inch, 2*inch])
        op_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(op_table)
        elements.append(Spacer(1, 20))
    
    # Add main table with detailed data
    elements.append(Paragraph("Detailed Report", heading_style))
    elements.append(Spacer(1, 10))
    
    # Create main table
    if queryset:
        # Get data from queryset
        data = []
        
        # Add header row
        header_row = []
        for col in columns:
            # Format column name (e.g., 'operation_type' becomes 'Operation Type')
            header = col.replace('_', ' ').title()
            header_row.append(header)
        data.append(header_row)
        
        # Add data rows
        for item in queryset:
            row = []
            for col in columns:
                # Handle different types of querysets (Rake vs Operation)
                if hasattr(item, 'get_' + col + '_display'):
                    # Use get_field_display for choice fields
                    value = getattr(item, 'get_' + col + '_display')()
                elif hasattr(item, col):
                    # Access regular fields
                    value = getattr(item, col)
                    
                    # Format value based on type
                    if isinstance(value, datetime):
                        # Convert to Asia/Kolkata timezone to match what's displayed in the UI
                        if value.tzinfo is None:
                            # If naive datetime, assume it's in UTC and convert to IST
                            ist = pytz.timezone('Asia/Kolkata')
                            value = pytz.utc.localize(value).astimezone(ist)
                        else:
                            # If already timezone aware, just convert to IST
                            ist = pytz.timezone('Asia/Kolkata')
                            value = value.astimezone(ist)
                        value = value.strftime('%d-%b-%Y %H:%M')
                    elif col == 'material' and hasattr(value, 'name'):
                        value = value.name
                else:
                    value = ''
                row.append(str(value))
            data.append(row)
        
        # Calculate column widths
        col_widths = []
        for col in columns:
            if col in ['date', 'shift', 'area']:
                col_widths.append(1 * inch)
            elif col in ['operation_type', 'tippler', 'rake_status']:
                col_widths.append(1.2 * inch)
            elif col in ['material', 'rake_material', 'rake_type', 'rake_id']:
                col_widths.append(1.5 * inch)
            elif col in ['quantity']:
                col_widths.append(0.8 * inch)
            elif col in ['rake_in_time', 'rake_completed_time']:
                col_widths.append(1.8 * inch)
            else:
                col_widths.append(1 * inch)
        
        # Create table with data
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Add table style
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),  # Center align all data cells
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Add explicit cell padding to prevent text from touching borders
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            # Ensure column separation
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
        ])
        table.setStyle(table_style)
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response

def get_tonnage_summary(queryset=None, **filters):
    """Get summary of tonnage by operation type"""
    from .models import Operation
    
    if queryset is None:
        queryset = Operation.objects.filter(**filters)
    
    summary = queryset.values('operation_type').annotate(
        total=Sum('quantity')
    ).order_by('operation_type')
    
    # Convert to dictionary for easier use in templates
    result = {}
    for item in summary:
        result[item['operation_type']] = item['total']
    
    return result

def get_material_summary(queryset=None, **filters):
    """Get summary of material tonnage"""
    from .models import Operation
    
    if queryset is None:
        queryset = Operation.objects.filter(**filters)
    
    # Material summary - using related names for FeedingOperation and others
    feeding_summary = queryset.filter(operation_type='feeding').values('material__name').annotate(
        total=Sum('quantity')
    ).order_by('-total')
    
    return feeding_summary 